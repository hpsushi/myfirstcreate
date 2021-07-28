# -*- coding: utf-8 -*-
"""
作者：Haiping.hu
日期：2021/7/27 8:50 上午
"""

import  requests
import openpyxl

def readdata():
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    max_row = sheet.max_row
    #print(max_row)
    case_list = []
    for i in range(2,max_row+1):
        switch = sheet.cell(row=i,column=9).value
        if switch ==1:
            dict1=dict(exexuteno=sheet.cell(row=i,column=1).value,
                       caseid = sheet.cell(row=i,column=2).value,
                       casename = sheet.cell(row=i,column=3).value,
                       method = sheet.cell(row=i,column=5).value,
                       url = sheet.cell(row=i,column=6).value,
                       data = sheet.cell(row=i,column=7).value,
                       expect = sheet.cell(row=i,column=8).value,
                       mt =sheet.cell(row=i,column=9).value,
        )
            case_list.append(dict1)
    return case_list




